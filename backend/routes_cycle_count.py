from flask import request, jsonify, session
from models import db, Tool, Checkout, AuditLog, Chemical
from models_cycle_count import (
    CycleCountSchedule, CycleCountBatch, CycleCountItem,
    CycleCountResult, CycleCountAdjustment
)
from datetime import datetime
from functools import wraps
import random
import logging
from utils.validation import validate_schema

# Helper function to create cycle count notifications
def create_cycle_count_notification(notification_type, message, batch_id=None, schedule_id=None, created_by=None):
    """Create a notification for cycle count activities"""
    try:
        # For now, we'll create audit log entries as notifications
        # In a full implementation, this would integrate with a proper notification system
        log = AuditLog(
            action_type=f'cycle_count_notification_{notification_type}',
            action_details=message,
            user_id=created_by
        )
        db.session.add(log)
        db.session.commit()

        # TODO: Integrate with email/push notification system
        # TODO: Create in-app notifications for relevant users

    except Exception as e:
        # Log the error but don't fail the main operation
        logging.error(f"Error creating cycle count notification: {str(e)}")
        # Consider whether to re-raise or continue silently

# Helper function to generate cycle count items for a batch
def generate_batch_items(batch_id, data):
    """
    Generate cycle count items for a batch based on the specified method
    """
    batch = CycleCountBatch.query.get(batch_id)
    if not batch:
        raise ValueError(f"Batch with ID {batch_id} not found")

    # Get method from schedule or data
    method = None
    if batch.schedule_id:
        schedule = CycleCountSchedule.query.get(batch.schedule_id)
        if schedule:
            method = schedule.method

    # Override method if provided in data
    if 'method' in data:
        method = data['method']

    if not method:
        raise ValueError("No method specified for generating items")

    # Get item types to include
    item_types = data.get('item_types', ['tool', 'chemical'])

    # Get filters
    location = data.get('location')
    category = data.get('category')

    # Get sample size or percentage
    sample_size = data.get('sample_size')
    sample_percentage = data.get('sample_percentage')

    # Generate items based on method
    items_to_add = []

    # Process tools
    if 'tool' in item_types:
        # Build query
        tool_query = Tool.query

        # Apply filters
        if location:
            tool_query = tool_query.filter(Tool.location == location)
        if category:
            tool_query = tool_query.filter(Tool.category == category)

        # Get all matching tools
        tools = tool_query.all()

        # Apply sampling if specified
        if method == 'ABC':
            # ABC analysis - prioritize high-value items
            # For tools, we'll use the most frequently checked out tools as "high value"
            tool_usage = {}
            for tool in tools:
                checkout_count = Checkout.query.filter_by(tool_id=tool.id).count()
                tool_usage[tool.id] = checkout_count

            # Sort tools by usage (descending)
            sorted_tools = sorted(tools, key=lambda t: tool_usage.get(t.id, 0), reverse=True)

            # Take top percentage as A, next as B, rest as C
            a_count = int(len(sorted_tools) * 0.2)  # Top 20% as A
            b_count = int(len(sorted_tools) * 0.3)  # Next 30% as B

            # A items - include all
            for tool in sorted_tools[:a_count]:
                items_to_add.append({
                    'batch_id': batch_id,
                    'item_type': 'tool',
                    'item_id': tool.id,
                    'expected_quantity': 1,
                    'expected_location': tool.location,
                    'status': 'pending'
                })

            # B items - include 50%
            b_tools = sorted_tools[a_count:a_count+b_count]
            if b_tools:
                sample_size = max(1, int(len(b_tools) * 0.5))
                b_sample = random.sample(b_tools, min(sample_size, len(b_tools)))
            else:
                b_sample = []
            for tool in b_sample:
                items_to_add.append({
                    'batch_id': batch_id,
                    'item_type': 'tool',
                    'item_id': tool.id,
                    'expected_quantity': 1,
                    'expected_location': tool.location,
                    'status': 'pending'
                })

            # C items - include 20%
            c_tools = sorted_tools[a_count+b_count:]
            if c_tools:
                sample_size = max(1, int(len(c_tools) * 0.2))
                c_sample = random.sample(c_tools, min(sample_size, len(c_tools)))
            else:
                c_sample = []
            for tool in c_sample:
                items_to_add.append({
                    'batch_id': batch_id,
                    'item_type': 'tool',
                    'item_id': tool.id,
                    'expected_quantity': 1,
                    'expected_location': tool.location,
                    'status': 'pending'
                })

        elif method == 'random':
            # Random sampling
            if len(tools) == 0:
                sampled_tools = []
            elif sample_size and sample_size < len(tools):
                # Take a random sample of specified size
                sampled_tools = random.sample(tools, sample_size)
            elif sample_percentage:
                # Take a random sample of specified percentage
                sample_count = max(1, int(len(tools) * sample_percentage / 100))
                sampled_tools = random.sample(tools, min(sample_count, len(tools)))
            else:
                # Include all tools
                sampled_tools = tools

            # Add sampled tools to items
            for tool in sampled_tools:
                items_to_add.append({
                    'batch_id': batch_id,
                    'item_type': 'tool',
                    'item_id': tool.id,
                    'expected_quantity': 1,
                    'expected_location': tool.location,
                    'status': 'pending'
                })

        elif method == 'location':
            # Group by location
            for tool in tools:
                items_to_add.append({
                    'batch_id': batch_id,
                    'item_type': 'tool',
                    'item_id': tool.id,
                    'expected_quantity': 1,
                    'expected_location': tool.location,
                    'status': 'pending'
                })

        elif method == 'category':
            # Group by category
            for tool in tools:
                items_to_add.append({
                    'batch_id': batch_id,
                    'item_type': 'tool',
                    'item_id': tool.id,
                    'expected_quantity': 1,
                    'expected_location': tool.location,
                    'status': 'pending'
                })

    # Process chemicals
    if 'chemical' in item_types:
        # Build query
        chemical_query = Chemical.query

        # Apply filters
        if location:
            chemical_query = chemical_query.filter(Chemical.location == location)
        if category:
            chemical_query = chemical_query.filter(Chemical.category == category)

        # Get all matching chemicals
        chemicals = chemical_query.all()

        # Apply sampling if specified
        if method == 'ABC':
            # ABC analysis - prioritize high-value items
            # For chemicals, we'll use quantity as a proxy for value
            sorted_chemicals = sorted(chemicals, key=lambda c: c.quantity, reverse=True)

            # Take top percentage as A, next as B, rest as C
            a_count = int(len(sorted_chemicals) * 0.2)  # Top 20% as A
            b_count = int(len(sorted_chemicals) * 0.3)  # Next 30% as B

            # A items - include all
            for chemical in sorted_chemicals[:a_count]:
                items_to_add.append({
                    'batch_id': batch_id,
                    'item_type': 'chemical',
                    'item_id': chemical.id,
                    'expected_quantity': chemical.quantity,
                    'expected_location': chemical.location,
                    'status': 'pending'
                })

            # B items - include 50%
            b_chemicals = sorted_chemicals[a_count:a_count+b_count]
            if b_chemicals:
                sample_size = max(1, int(len(b_chemicals) * 0.5))
                b_sample = random.sample(b_chemicals, min(sample_size, len(b_chemicals)))
            else:
                b_sample = []
            for chemical in b_sample:
                items_to_add.append({
                    'batch_id': batch_id,
                    'item_type': 'chemical',
                    'item_id': chemical.id,
                    'expected_quantity': chemical.quantity,
                    'expected_location': chemical.location,
                    'status': 'pending'
                })

            # C items - include 20%
            c_chemicals = sorted_chemicals[a_count+b_count:]
            if c_chemicals:
                sample_size = max(1, int(len(c_chemicals) * 0.2))
                c_sample = random.sample(c_chemicals, min(sample_size, len(c_chemicals)))
            else:
                c_sample = []
            for chemical in c_sample:
                items_to_add.append({
                    'batch_id': batch_id,
                    'item_type': 'chemical',
                    'item_id': chemical.id,
                    'expected_quantity': chemical.quantity,
                    'expected_location': chemical.location,
                    'status': 'pending'
                })

        elif method == 'random':
            # Random sampling
            if len(chemicals) == 0:
                sampled_chemicals = []
            elif sample_size and sample_size < len(chemicals):
                # Take a random sample of specified size
                sampled_chemicals = random.sample(chemicals, sample_size)
            elif sample_percentage:
                # Take a random sample of specified percentage
                sample_count = max(1, int(len(chemicals) * sample_percentage / 100))
                sampled_chemicals = random.sample(chemicals, min(sample_count, len(chemicals)))
            else:
                # Include all chemicals
                sampled_chemicals = chemicals

            # Add sampled chemicals to items
            for chemical in sampled_chemicals:
                items_to_add.append({
                    'batch_id': batch_id,
                    'item_type': 'chemical',
                    'item_id': chemical.id,
                    'expected_quantity': chemical.quantity,
                    'expected_location': chemical.location,
                    'status': 'pending'
                })

        elif method in ['location', 'category']:
            # Include all chemicals
            for chemical in chemicals:
                items_to_add.append({
                    'batch_id': batch_id,
                    'item_type': 'chemical',
                    'item_id': chemical.id,
                    'expected_quantity': chemical.quantity,
                    'expected_location': chemical.location,
                    'status': 'pending'
                })

    # Add all items to database
    for item_data in items_to_add:
        item = CycleCountItem(**item_data)
        db.session.add(item)

    db.session.commit()

    return len(items_to_add)

def update_item_from_count_result(cycle_count_item, count_result):
    """Update the actual tool or chemical based on cycle count results"""
    try:
        if cycle_count_item.item_type == 'tool':
            # Update tool
            tool = Tool.query.get(cycle_count_item.item_id)
            if tool:
                # Update location if different
                if count_result.actual_location and count_result.actual_location != tool.location:
                    old_location = tool.location
                    tool.location = count_result.actual_location

                    # Log location change
                    log = AuditLog(
                        action_type='tool_location_updated_from_cycle_count',
                        action_details=f"Tool {tool.part_number} location updated from '{old_location}' to '{count_result.actual_location}' based on cycle count"
                    )
                    db.session.add(log)

                # Update condition if provided
                if count_result.condition and count_result.condition != tool.condition:
                    old_condition = tool.condition
                    tool.condition = count_result.condition

                    # Log condition change
                    log = AuditLog(
                        action_type='tool_condition_updated_from_cycle_count',
                        action_details=f"Tool {tool.part_number} condition updated from '{old_condition}' to '{count_result.condition}' based on cycle count"
                    )
                    db.session.add(log)

                db.session.commit()

        elif cycle_count_item.item_type == 'chemical':
            # Update chemical
            chemical = Chemical.query.get(cycle_count_item.item_id)
            if chemical:
                # Update quantity if different
                if count_result.actual_quantity != chemical.quantity:
                    old_quantity = chemical.quantity
                    chemical.quantity = count_result.actual_quantity

                    # Update status based on new quantity
                    if chemical.quantity <= 0:
                        chemical.status = 'out_of_stock'
                    elif chemical.is_low_stock():
                        chemical.status = 'low_stock'
                    else:
                        chemical.status = 'available'

                    # Log quantity change
                    log = AuditLog(
                        action_type='chemical_quantity_updated_from_cycle_count',
                        action_details=f"Chemical {chemical.part_number} - {chemical.lot_number} quantity updated from {old_quantity} to {count_result.actual_quantity} based on cycle count"
                    )
                    db.session.add(log)

                # Update location if different
                if count_result.actual_location and count_result.actual_location != chemical.location:
                    old_location = chemical.location
                    chemical.location = count_result.actual_location

                    # Log location change
                    log = AuditLog(
                        action_type='chemical_location_updated_from_cycle_count',
                        action_details=f"Chemical {chemical.part_number} - {chemical.lot_number} location updated from '{old_location}' to '{count_result.actual_location}' based on cycle count"
                    )
                    db.session.add(log)

                db.session.commit()

    except Exception as e:
        print(f"Error updating item from count result: {str(e)}")
        # Don't raise the exception to avoid breaking the count submission

def tool_manager_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Authentication required'}), 401

        # Allow access for admins or Materials department users
        if session.get('is_admin', False) or session.get('department') == 'Materials':
            return f(*args, **kwargs)

        return jsonify({'error': 'Tool management privileges required'}), 403
    return decorated_function

def register_cycle_count_routes(app):
    # Get all cycle count schedules
    @app.route('/api/cycle-counts/schedules', methods=['GET'])
    @tool_manager_required
    def get_cycle_count_schedules():
        try:
            # Get query parameters
            active_only = request.args.get('active_only', 'false').lower() == 'true'

            # Build query
            query = CycleCountSchedule.query

            # Filter by active status if requested
            if active_only:
                query = query.filter_by(is_active=True)

            # Execute query
            schedules = query.order_by(CycleCountSchedule.created_at.desc()).all()

            # Return results
            return jsonify([schedule.to_dict() for schedule in schedules]), 200

        except Exception as e:
            print(f"Error getting cycle count schedules: {str(e)}")
            return jsonify({'error': 'An error occurred while fetching cycle count schedules'}), 500

    # Get a specific cycle count schedule
    @app.route('/api/cycle-counts/schedules/<int:id>', methods=['GET'])
    @tool_manager_required
    def get_cycle_count_schedule(id):
        try:
            # Get query parameters
            include_batches = request.args.get('include_batches', 'false').lower() == 'true'

            # Get schedule
            schedule = CycleCountSchedule.query.get_or_404(id)

            # Return result
            return jsonify(schedule.to_dict(include_batches=include_batches)), 200

        except Exception as e:
            print(f"Error getting cycle count schedule {id}: {str(e)}")
            return jsonify({'error': f'An error occurred while fetching cycle count schedule {id}'}), 500

    # Create a new cycle count schedule
    @app.route('/api/cycle-counts/schedules', methods=['POST'])
    @tool_manager_required
    def create_cycle_count_schedule():
        try:
            # Get request data
            data = request.get_json()

            # Validate data using schema
            validated_data = validate_schema(data, 'cycle_count_schedule')

            # Create new schedule
            schedule = CycleCountSchedule(
                name=validated_data['name'],
                description=validated_data.get('description', ''),
                frequency=validated_data['frequency'],
                method=validated_data['method'],
                created_by=session['user_id'],
                is_active=validated_data.get('is_active', True)
            )

            # Save to database
            db.session.add(schedule)
            db.session.commit()

            # Log the action
            log = AuditLog(
                action_type='cycle_count_schedule_created',
                action_details=f"Cycle count schedule '{data['name']}' created"
            )
            db.session.add(log)
            db.session.commit()

            # Return result
            return jsonify(schedule.to_dict()), 201

        except Exception as e:
            db.session.rollback()
            print(f"Error creating cycle count schedule: {str(e)}")
            return jsonify({'error': 'An error occurred while creating cycle count schedule'}), 500

    # Update a cycle count schedule
    @app.route('/api/cycle-counts/schedules/<int:id>', methods=['PUT'])
    @tool_manager_required
    def update_cycle_count_schedule(id):
        try:
            # Get request data
            data = request.get_json()

            # Get schedule
            schedule = CycleCountSchedule.query.get_or_404(id)

            # Update fields
            if 'name' in data:
                schedule.name = data['name']
            if 'description' in data:
                schedule.description = data['description']
            if 'frequency' in data:
                schedule.frequency = data['frequency']
            if 'method' in data:
                schedule.method = data['method']
            if 'is_active' in data:
                schedule.is_active = data['is_active']

            # Save to database
            db.session.commit()

            # Log the action
            log = AuditLog(
                action_type='cycle_count_schedule_updated',
                action_details=f"Cycle count schedule '{schedule.name}' updated"
            )
            db.session.add(log)
            db.session.commit()

            # Return result
            return jsonify(schedule.to_dict()), 200

        except Exception as e:
            db.session.rollback()
            print(f"Error updating cycle count schedule {id}: {str(e)}")
            return jsonify({'error': f'An error occurred while updating cycle count schedule {id}'}), 500

    # Delete a cycle count schedule
    @app.route('/api/cycle-counts/schedules/<int:id>', methods=['DELETE'])
    @tool_manager_required
    def delete_cycle_count_schedule(id):
        try:
            # Get schedule
            schedule = CycleCountSchedule.query.get_or_404(id)

            # Check if schedule has batches
            if schedule.batches:
                # Don't delete, just mark as inactive
                schedule.is_active = False
                db.session.commit()

                # Log the action
                log = AuditLog(
                    action_type='cycle_count_schedule_deactivated',
                    action_details=f"Cycle count schedule '{schedule.name}' deactivated"
                )
                db.session.add(log)
                db.session.commit()

                return jsonify({'message': f"Schedule '{schedule.name}' has batches and cannot be deleted. It has been deactivated instead."}), 200

            # Delete schedule
            db.session.delete(schedule)

            # Log the action
            log = AuditLog(
                action_type='cycle_count_schedule_deleted',
                action_details=f"Cycle count schedule '{schedule.name}' deleted"
            )
            db.session.add(log)
            db.session.commit()

            # Return result
            return jsonify({'message': f"Schedule '{schedule.name}' deleted successfully"}), 200

        except Exception as e:
            db.session.rollback()
            print(f"Error deleting cycle count schedule {id}: {str(e)}")
            return jsonify({'error': f'An error occurred while deleting cycle count schedule {id}'}), 500

    # Get all cycle count batches
    @app.route('/api/cycle-counts/batches', methods=['GET'])
    @tool_manager_required
    def get_cycle_count_batches():
        try:
            # Get query parameters
            status = request.args.get('status')
            schedule_id = request.args.get('schedule_id', type=int)

            # Build query
            query = CycleCountBatch.query

            # Apply filters
            if status:
                query = query.filter_by(status=status)
            if schedule_id:
                query = query.filter_by(schedule_id=schedule_id)

            # Execute query
            batches = query.order_by(CycleCountBatch.created_at.desc()).all()

            # Return results
            return jsonify([batch.to_dict() for batch in batches]), 200

        except Exception as e:
            print(f"Error getting cycle count batches: {str(e)}")
            return jsonify({'error': 'An error occurred while fetching cycle count batches'}), 500

    # Get a specific cycle count batch
    @app.route('/api/cycle-counts/batches/<int:id>', methods=['GET'])
    @tool_manager_required
    def get_cycle_count_batch(id):
        try:
            # Get query parameters
            include_items = request.args.get('include_items', 'false').lower() == 'true'

            # Get batch
            batch = CycleCountBatch.query.get_or_404(id)

            # Return result
            return jsonify(batch.to_dict(include_items=include_items)), 200

        except Exception as e:
            print(f"Error getting cycle count batch {id}: {str(e)}")
            return jsonify({'error': f'An error occurred while fetching cycle count batch {id}'}), 500

    # Create a new cycle count batch
    @app.route('/api/cycle-counts/batches', methods=['POST'])
    @tool_manager_required
    def create_cycle_count_batch():
        try:
            # Get request data
            data = request.get_json()

            # Validate data using schema
            validated_data = validate_schema(data, 'cycle_count_batch')

            # Validate dates
            if 'start_date' in validated_data and 'end_date' in validated_data and validated_data['start_date'] and validated_data['end_date']:
                start_date = datetime.fromisoformat(validated_data['start_date'])
                end_date = datetime.fromisoformat(validated_data['end_date'])
                if end_date < start_date:
                    return jsonify({'error': 'End date cannot be before start date'}), 400

            # Create new batch
            batch = CycleCountBatch(
                schedule_id=validated_data.get('schedule_id'),
                name=validated_data['name'],
                status='pending',
                start_date=datetime.fromisoformat(validated_data['start_date']) if validated_data.get('start_date') else None,
                end_date=datetime.fromisoformat(validated_data['end_date']) if validated_data.get('end_date') else None,
                created_by=session['user_id'],
                notes=validated_data.get('notes', '')
            )

            # Save to database
            db.session.add(batch)
            db.session.commit()

            # Generate items for the batch if requested
            if validated_data.get('generate_items', False):
                generate_batch_items(batch.id, validated_data)

            # Log the action
            log = AuditLog(
                action_type='cycle_count_batch_created',
                action_details=f"Cycle count batch '{validated_data['name']}' created"
            )
            db.session.add(log)
            db.session.commit()

            # Create notification for batch creation
            create_cycle_count_notification(
                'batch_created',
                f"New cycle count batch '{validated_data['name']}' has been created",
                batch_id=batch.id,
                created_by=session['user_id']
            )

            # Return result
            return jsonify(batch.to_dict()), 201

        except Exception as e:
            db.session.rollback()
            print(f"Error creating cycle count batch: {str(e)}")
            return jsonify({'error': 'An error occurred while creating cycle count batch'}), 500

    # Update a cycle count batch
    @app.route('/api/cycle-counts/batches/<int:id>', methods=['PUT'])
    @tool_manager_required
    def update_cycle_count_batch(id):
        try:
            # Get request data
            data = request.get_json()

            # Get batch
            batch = CycleCountBatch.query.get_or_404(id)

            # Validate dates
            start_date = None
            end_date = None

            if 'start_date' in data and data['start_date']:
                start_date = datetime.fromisoformat(data['start_date'])
            else:
                start_date = batch.start_date

            if 'end_date' in data and data['end_date']:
                end_date = datetime.fromisoformat(data['end_date'])
            else:
                end_date = batch.end_date

            if start_date and end_date and end_date < start_date:
                return jsonify({'error': 'End date cannot be before start date'}), 400

            # Update fields
            if 'name' in data:
                batch.name = data['name']
            if 'status' in data:
                batch.status = data['status']
            if 'schedule_id' in data:
                batch.schedule_id = data['schedule_id']
            if 'start_date' in data:
                batch.start_date = start_date
            if 'end_date' in data:
                batch.end_date = end_date
            if 'notes' in data:
                batch.notes = data['notes']

            # Save to database
            db.session.commit()

            # Log the action
            log = AuditLog(
                action_type='cycle_count_batch_updated',
                action_details=f"Cycle count batch '{batch.name}' updated"
            )
            db.session.add(log)
            db.session.commit()

            # Return result
            return jsonify(batch.to_dict()), 200

        except Exception as e:
            db.session.rollback()
            print(f"Error updating cycle count batch {id}: {str(e)}")
            return jsonify({'error': f'An error occurred while updating cycle count batch {id}'}), 500

    # Delete a cycle count batch
    @app.route('/api/cycle-counts/batches/<int:id>', methods=['DELETE'])
    @tool_manager_required
    def delete_cycle_count_batch(id):
        try:
            # Get batch
            batch = CycleCountBatch.query.get_or_404(id)

            # Check if batch has items with results
            has_results = False
            for item in batch.items:
                if item.results:
                    has_results = True
                    break

            if has_results:
                # Don't delete, just mark as cancelled
                batch.status = 'cancelled'
                db.session.commit()

                # Log the action
                log = AuditLog(
                    action_type='cycle_count_batch_cancelled',
                    action_details=f"Cycle count batch '{batch.name}' cancelled"
                )
                db.session.add(log)
                db.session.commit()

                return jsonify({'message': f"Batch '{batch.name}' has count results and cannot be deleted. It has been cancelled instead."}), 200

            # Delete batch items
            for item in batch.items:
                db.session.delete(item)

            # Delete batch
            db.session.delete(batch)

            # Log the action
            log = AuditLog(
                action_type='cycle_count_batch_deleted',
                action_details=f"Cycle count batch '{batch.name}' deleted"
            )
            db.session.add(log)
            db.session.commit()

            # Return result
            return jsonify({'message': f"Batch '{batch.name}' deleted successfully"}), 200

        except Exception as e:
            db.session.rollback()
            print(f"Error deleting cycle count batch {id}: {str(e)}")
            return jsonify({'error': f'An error occurred while deleting cycle count batch {id}'}), 500

    # Get all cycle count items for a batch
    @app.route('/api/cycle-counts/batches/<int:batch_id>/items', methods=['GET'])
    @tool_manager_required
    def get_cycle_count_items(batch_id):
        try:
            # Get query parameters
            status = request.args.get('status')
            assigned_to = request.args.get('assigned_to')
            item_type = request.args.get('item_type')

            # Build query
            query = CycleCountItem.query.filter_by(batch_id=batch_id)

            # Apply filters
            if status:
                query = query.filter_by(status=status)
            if assigned_to:
                query = query.filter_by(assigned_to=assigned_to)
            if item_type:
                query = query.filter_by(item_type=item_type)

            # Execute query
            items = query.all()

            # Return results
            return jsonify([item.to_dict() for item in items]), 200

        except Exception as e:
            print(f"Error getting cycle count items for batch {batch_id}: {str(e)}")
            return jsonify({'error': f'An error occurred while fetching cycle count items for batch {batch_id}'}), 500

    # Get a specific cycle count item
    @app.route('/api/cycle-counts/items/<int:id>', methods=['GET'])
    @tool_manager_required
    def get_cycle_count_item(id):
        try:
            # Get query parameters
            include_results = request.args.get('include_results', 'false').lower() == 'true'

            # Get item
            item = CycleCountItem.query.get_or_404(id)

            # Return result
            return jsonify(item.to_dict(include_results=include_results)), 200

        except Exception as e:
            print(f"Error getting cycle count item {id}: {str(e)}")
            return jsonify({'error': f'An error occurred while fetching cycle count item {id}'}), 500

    # Update a cycle count item
    @app.route('/api/cycle-counts/items/<int:id>', methods=['PUT'])
    @tool_manager_required
    def update_cycle_count_item(id):
        try:
            # Get request data
            data = request.get_json()

            # Get item
            item = CycleCountItem.query.get_or_404(id)

            # Update fields
            if 'assigned_to' in data:
                item.assigned_to = data['assigned_to']
            if 'status' in data:
                item.status = data['status']

            # Save to database
            db.session.commit()

            # Log the action
            log = AuditLog(
                action_type='cycle_count_item_updated',
                action_details=f"Cycle count item {id} updated"
            )
            db.session.add(log)
            db.session.commit()

            # Return result
            return jsonify(item.to_dict()), 200

        except Exception as e:
            db.session.rollback()
            print(f"Error updating cycle count item {id}: {str(e)}")
            return jsonify({'error': f'An error occurred while updating cycle count item {id}'}), 500

    # Submit a count result for an item
    @app.route('/api/cycle-counts/items/<int:item_id>/count', methods=['POST'])
    @tool_manager_required
    def submit_count_result(item_id):
        try:
            # Get request data
            data = request.get_json()

            # Validate data using schema
            validated_data = validate_schema(data, 'cycle_count_result')

            # Get item
            item = CycleCountItem.query.get_or_404(item_id)

            # Check if item is already counted
            if item.status == 'counted':
                return jsonify({'error': f'Item {item_id} has already been counted'}), 400

            # Determine if there's a discrepancy
            has_discrepancy = False
            discrepancy_type = None
            discrepancy_notes = []

            # Check quantity discrepancy
            if item.expected_quantity != validated_data['actual_quantity']:
                has_discrepancy = True
                discrepancy_type = 'quantity'
                discrepancy_notes.append(f"Expected quantity: {item.expected_quantity}, Actual: {validated_data['actual_quantity']}")

            # Check location discrepancy
            if 'actual_location' in validated_data and item.expected_location != validated_data['actual_location']:
                has_discrepancy = True
                discrepancy_type = discrepancy_type or 'location'
                discrepancy_notes.append(f"Expected location: {item.expected_location}, Actual: {validated_data['actual_location']}")

            # Create count result
            result = CycleCountResult(
                item_id=item_id,
                counted_by=session['user_id'],
                actual_quantity=validated_data['actual_quantity'],
                actual_location=validated_data.get('actual_location'),
                condition=validated_data.get('condition'),
                notes=validated_data.get('notes', ''),
                has_discrepancy=has_discrepancy,
                discrepancy_type=discrepancy_type,
                discrepancy_notes='\n'.join(discrepancy_notes) if discrepancy_notes else None
            )

            # Update item status
            item.status = 'counted'

            # Save to database
            db.session.add(result)
            db.session.commit()

            # Update the actual item based on count results if there are discrepancies
            if has_discrepancy:
                update_item_from_count_result(item, result)

            # Log the action
            log = AuditLog(
                action_type='cycle_count_result_submitted',
                action_details=f"Count result submitted for item {item_id}"
            )
            db.session.add(log)
            db.session.commit()

            # Create notification for discrepancies
            if has_discrepancy:
                create_cycle_count_notification(
                    'discrepancy_found',
                    f"Discrepancy found in item {item.item_number}: {discrepancy_type}",
                    batch_id=item.batch_id,
                    created_by=session['user_id']
                )

            # Return result
            return jsonify(result.to_dict()), 201

        except Exception as e:
            db.session.rollback()
            print(f"Error submitting count result for item {item_id}: {str(e)}")
            return jsonify({'error': f'An error occurred while submitting count result for item {item_id}'}), 500

    # Get all count results with discrepancies
    @app.route('/api/cycle-counts/discrepancies', methods=['GET'])
    @tool_manager_required
    def get_count_discrepancies():
        try:
            # Get query parameters
            batch_id = request.args.get('batch_id')

            # Build query
            query = CycleCountResult.query.filter_by(has_discrepancy=True)

            # Apply batch filter if provided
            if batch_id:
                query = query.join(CycleCountItem).filter(CycleCountItem.batch_id == batch_id)

            # Execute query
            results = query.all()

            # Return results
            return jsonify([result.to_dict(include_item=True) for result in results]), 200

        except Exception as e:
            print(f"Error getting count discrepancies: {str(e)}")
            return jsonify({'error': 'An error occurred while fetching count discrepancies'}), 500

    # Get a specific count result
    @app.route('/api/cycle-counts/results/<int:result_id>', methods=['GET'])
    @tool_manager_required
    def get_count_result(result_id):
        try:
            # Get result
            result = CycleCountResult.query.get_or_404(result_id)

            # The item details are included via include_item=True

            # Return result with item details
            result_dict = result.to_dict(include_item=True, include_adjustments=True)

            return jsonify(result_dict), 200

        except Exception as e:
            print(f"Error getting count result {result_id}: {str(e)}")
            return jsonify({'error': f'An error occurred while fetching count result {result_id}'}), 500

    # Approve and process a count adjustment
    @app.route('/api/cycle-counts/results/<int:result_id>/adjust', methods=['POST'])
    @tool_manager_required
    def approve_count_adjustment(result_id):
        try:
            # Get request data
            data = request.get_json()

            # Validate required fields
            if not all(key in data for key in ['adjustment_type', 'new_value']):
                return jsonify({'error': 'Missing required fields: adjustment_type, new_value'}), 400

            # Get result
            result = CycleCountResult.query.get_or_404(result_id)

            # Get item
            item = CycleCountItem.query.get_or_404(result.item_id)

            # Get old value based on adjustment type
            old_value = None
            if data['adjustment_type'] == 'quantity':
                if item.item_type == 'tool':
                    old_value = '1'  # Tools always have quantity 1
                elif item.item_type == 'chemical':
                    chemical = Chemical.query.get(item.item_id)
                    if chemical:
                        old_value = str(chemical.quantity)
            elif data['adjustment_type'] == 'location':
                if item.item_type == 'tool':
                    tool = Tool.query.get(item.item_id)
                    if tool:
                        old_value = tool.location
                elif item.item_type == 'chemical':
                    chemical = Chemical.query.get(item.item_id)
                    if chemical:
                        old_value = chemical.location
            elif data['adjustment_type'] == 'condition':
                if item.item_type == 'tool':
                    tool = Tool.query.get(item.item_id)
                    if tool:
                        old_value = tool.condition
            elif data['adjustment_type'] == 'status':
                if item.item_type == 'tool':
                    tool = Tool.query.get(item.item_id)
                    if tool:
                        old_value = tool.status
                elif item.item_type == 'chemical':
                    chemical = Chemical.query.get(item.item_id)
                    if chemical:
                        old_value = chemical.status

            # Create adjustment record
            adjustment = CycleCountAdjustment(
                result_id=result_id,
                approved_by=session['user_id'],
                adjustment_type=data['adjustment_type'],
                old_value=old_value,
                new_value=data['new_value'],
                notes=data.get('notes', '')
            )

            # Apply the adjustment to the inventory
            if item.item_type == 'tool':
                tool = Tool.query.get(item.item_id)
                if tool:
                    if data['adjustment_type'] == 'location':
                        tool.location = data['new_value']
                    elif data['adjustment_type'] == 'condition':
                        tool.condition = data['new_value']
                    elif data['adjustment_type'] == 'status':
                        tool.status = data['new_value']
                        if data['new_value'] in ['maintenance', 'retired']:
                            tool.status_reason = data.get('notes', 'Updated from cycle count')
            elif item.item_type == 'chemical':
                chemical = Chemical.query.get(item.item_id)
                if chemical:
                    if data['adjustment_type'] == 'quantity':
                        chemical.quantity = float(data['new_value'])
                    elif data['adjustment_type'] == 'location':
                        chemical.location = data['new_value']
                    elif data['adjustment_type'] == 'status':
                        chemical.status = data['new_value']

            # Save to database
            db.session.add(adjustment)
            db.session.commit()

            # Log the action
            log = AuditLog(
                action_type='cycle_count_adjustment_approved',
                action_details=f"Count adjustment approved for result {result_id}"
            )
            db.session.add(log)
            db.session.commit()

            # Return result
            return jsonify(adjustment.to_dict()), 201

        except Exception as e:
            db.session.rollback()
            print(f"Error approving count adjustment for result {result_id}: {str(e)}")
            return jsonify({'error': f'An error occurred while approving count adjustment for result {result_id}'}), 500

    # Get cycle count statistics
    @app.route('/api/cycle-counts/stats', methods=['GET'])
    @tool_manager_required
    def get_cycle_count_stats():
        try:
            # Get overall statistics
            total_schedules = CycleCountSchedule.query.count()
            active_schedules = CycleCountSchedule.query.filter_by(is_active=True).count()

            total_batches = CycleCountBatch.query.count()
            pending_batches = CycleCountBatch.query.filter_by(status='pending').count()
            in_progress_batches = CycleCountBatch.query.filter_by(status='in_progress').count()
            completed_batches = CycleCountBatch.query.filter_by(status='completed').count()

            total_items = CycleCountItem.query.count()
            counted_items = CycleCountItem.query.filter_by(status='counted').count()
            pending_items = CycleCountItem.query.filter_by(status='pending').count()

            total_results = CycleCountResult.query.count()
            discrepancy_results = CycleCountResult.query.filter_by(has_discrepancy=True).count()

            # Calculate accuracy rate
            accuracy_rate = 0
            if total_results > 0:
                accuracy_rate = round(100 * (total_results - discrepancy_results) / total_results, 2)

            # Get recent batches
            recent_batches = CycleCountBatch.query.order_by(CycleCountBatch.created_at.desc()).limit(5).all()

            # Return results
            return jsonify({
                'schedules': {
                    'total': total_schedules,
                    'active': active_schedules
                },
                'batches': {
                    'total': total_batches,
                    'pending': pending_batches,
                    'in_progress': in_progress_batches,
                    'completed': completed_batches
                },
                'items': {
                    'total': total_items,
                    'counted': counted_items,
                    'pending': pending_items,
                    'completion_rate': round(100 * counted_items / total_items, 2) if total_items > 0 else 0
                },
                'results': {
                    'total': total_results,
                    'with_discrepancies': discrepancy_results,
                    'accuracy_rate': accuracy_rate
                },
                'recent_batches': [batch.to_dict() for batch in recent_batches]
            }), 200

        except Exception as e:
            print(f"Error getting cycle count statistics: {str(e)}")
            return jsonify({'error': 'An error occurred while fetching cycle count statistics'}), 500

    # Export cycle count batch data
    @app.route('/api/cycle-counts/batches/<int:batch_id>/export', methods=['GET'])
    @tool_manager_required
    def export_cycle_count_batch(batch_id):
        try:
            from io import StringIO
            import csv
            from flask import make_response

            # Get format parameter
            format_type = request.args.get('format', 'csv').lower()

            # Get batch
            batch = CycleCountBatch.query.get_or_404(batch_id)

            # Get all items for the batch
            items = CycleCountItem.query.filter_by(batch_id=batch_id).all()

            if format_type == 'excel':
                # Excel export using openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                import io

                wb = Workbook()
                ws = wb.active
                ws.title = "Batch Items"

                # Add batch information
                ws['A1'] = f"Cycle Count Batch: {batch.name}"
                ws['A1'].font = Font(size=16, bold=True)

                ws['A3'] = f"Batch ID: {batch.id}"
                ws['A4'] = f"Status: {batch.status}"
                ws['A5'] = f"Created: {batch.created_at.strftime('%Y-%m-%d %H:%M:%S')}"
                ws['A6'] = f"Total Items: {len(items)}"

                # Headers
                headers = [
                    'Item ID', 'Item Type', 'Item Reference ID', 'Expected Quantity',
                    'Expected Location', 'Status', 'Assigned To'
                ]

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=8, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                # Add data
                for row, item in enumerate(items, 9):
                    ws.cell(row=row, column=1, value=item.id)
                    ws.cell(row=row, column=2, value=item.item_type)
                    ws.cell(row=row, column=3, value=item.item_id)
                    ws.cell(row=row, column=4, value=item.expected_quantity)
                    ws.cell(row=row, column=5, value=item.expected_location)
                    ws.cell(row=row, column=6, value=item.status)
                    ws.cell(row=row, column=7, value=item.assigned_to or '')

                # Save to buffer
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                # Create response
                response = make_response(buffer.getvalue())
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response.headers['Content-Disposition'] = f'attachment; filename=cycle_count_batch_{batch_id}.xlsx'

                return response

            else:
                # CSV export
                output = StringIO()
                writer = csv.writer(output)

                # Write header
                writer.writerow([
                    'Item ID', 'Item Type', 'Item Reference ID', 'Expected Quantity',
                    'Expected Location', 'Status', 'Assigned To'
                ])

                # Write data rows
                for item in items:
                    writer.writerow([
                        item.id,
                        item.item_type,
                        item.item_id,
                        item.expected_quantity,
                        item.expected_location,
                        item.status,
                        item.assigned_to or ''
                    ])

                # Create response
                response = make_response(output.getvalue())
                response.headers['Content-Type'] = 'text/csv'
                response.headers['Content-Disposition'] = f'attachment; filename=cycle_count_batch_{batch_id}.csv'

                return response

        except Exception as e:
            print(f"Error exporting cycle count batch {batch_id}: {str(e)}")
            return jsonify({'error': f'An error occurred while exporting cycle count batch {batch_id}'}), 500

    # Import cycle count results
    @app.route('/api/cycle-counts/batches/<int:batch_id>/import', methods=['POST'])
    @tool_manager_required
    def import_cycle_count_results(batch_id):
        try:
            import csv
            from io import StringIO

            # Get batch
            batch = CycleCountBatch.query.get_or_404(batch_id)

            # Get uploaded file
            if 'file' not in request.files:
                return jsonify({'error': 'No file uploaded'}), 400

            file = request.files['file']
            if file.filename == '':
                return jsonify({'error': 'No file selected'}), 400

            # Read CSV data
            csv_data = file.read().decode('utf-8')
            csv_reader = csv.DictReader(StringIO(csv_data))

            imported_count = 0
            errors = []

            for row_num, row in enumerate(csv_reader, start=2):
                try:
                    # Validate required fields
                    if not row.get('Item ID'):
                        errors.append(f"Row {row_num}: Missing Item ID")
                        continue

                    item_id = int(row['Item ID'])

                    # Get the cycle count item
                    item = CycleCountItem.query.filter_by(
                        id=item_id,
                        batch_id=batch_id
                    ).first()

                    if not item:
                        errors.append(f"Row {row_num}: Item ID {item_id} not found in batch")
                        continue

                    # Check if already counted
                    if item.status == 'counted':
                        errors.append(f"Row {row_num}: Item {item_id} already counted")
                        continue

                    # Get count data
                    actual_quantity = float(row.get('Actual Quantity', 0))
                    actual_location = row.get('Actual Location', '')
                    condition = row.get('Condition', '')
                    notes = row.get('Notes', '')

                    # Determine discrepancy
                    has_discrepancy = False
                    discrepancy_type = None
                    discrepancy_notes = []

                    if actual_quantity != item.expected_quantity:
                        has_discrepancy = True
                        discrepancy_type = 'quantity'
                        discrepancy_notes.append(f"Expected: {item.expected_quantity}, Actual: {actual_quantity}")

                    if actual_location != item.expected_location:
                        has_discrepancy = True
                        if not discrepancy_type:
                            discrepancy_type = 'location'
                        discrepancy_notes.append(f"Expected location: {item.expected_location}, Actual: {actual_location}")

                    # Create count result
                    result = CycleCountResult(
                        item_id=item_id,
                        counted_by=session['user_id'],
                        actual_quantity=actual_quantity,
                        actual_location=actual_location,
                        condition=condition,
                        notes=notes,
                        has_discrepancy=has_discrepancy,
                        discrepancy_type=discrepancy_type,
                        discrepancy_notes='\n'.join(discrepancy_notes) if discrepancy_notes else None
                    )

                    # Update item status
                    item.status = 'counted'

                    # Save to database
                    db.session.add(result)
                    imported_count += 1

                except ValueError as ve:
                    errors.append(f"Row {row_num}: Invalid data - {str(ve)}")
                except Exception as e:
                    errors.append(f"Row {row_num}: Error - {str(e)}")

            # Commit changes
            db.session.commit()

            # Log the action
            log = AuditLog(
                action_type='cycle_count_results_imported',
                action_details=f"Imported {imported_count} count results for batch {batch_id}"
            )
            db.session.add(log)
            db.session.commit()

            return jsonify({
                'message': f'Successfully imported {imported_count} count results',
                'imported_count': imported_count,
                'errors': errors
            }), 200

        except Exception as e:
            db.session.rollback()
            print(f"Error importing cycle count results for batch {batch_id}: {str(e)}")
            return jsonify({'error': 'An error occurred while importing cycle count results'}), 500

    # Get advanced cycle count analytics
    @app.route('/api/cycle-counts/analytics', methods=['GET'])
    @tool_manager_required
    def get_cycle_count_analytics():
        try:
            from datetime import datetime, timedelta
            from sqlalchemy import func, and_

            # Get query parameters
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')

            # Default to last 30 days if no dates provided
            if not start_date:
                start_date = (datetime.now() - timedelta(days=30)).isoformat()
            if not end_date:
                end_date = datetime.now().isoformat()

            start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))

            # Accuracy trends over time - using simpler approach for SQLite compatibility
            accuracy_trends_raw = db.session.query(
                func.date(CycleCountResult.counted_at).label('date'),
                func.count(CycleCountResult.id).label('total_counts'),
                func.count(CycleCountResult.id).filter(~CycleCountResult.has_discrepancy).label('accurate_counts')
            ).filter(
                and_(
                    CycleCountResult.counted_at >= start_dt,
                    CycleCountResult.counted_at <= end_dt
                )
            ).group_by(
                func.date(CycleCountResult.counted_at)
            ).order_by(
                func.date(CycleCountResult.counted_at)
            ).all()

            # Convert to list for easier handling
            accuracy_trends = []
            for trend in accuracy_trends_raw:
                accuracy_trends.append({
                    'date': trend.date,
                    'total_counts': trend.total_counts,
                    'accurate_counts': trend.accurate_counts or 0
                })

            # Discrepancy analysis by type
            discrepancy_types = db.session.query(
                CycleCountResult.discrepancy_type,
                func.count(CycleCountResult.id).label('count')
            ).filter(
                and_(
                    CycleCountResult.has_discrepancy,
                    CycleCountResult.counted_at >= start_dt,
                    CycleCountResult.counted_at <= end_dt
                )
            ).group_by(
                CycleCountResult.discrepancy_type
            ).all()

            # Performance by user - using simpler approach for SQLite compatibility
            user_performance_raw = db.session.query(
                CycleCountResult.counted_by,
                func.count(CycleCountResult.id).label('total_counts'),
                func.count(CycleCountResult.id).filter(~CycleCountResult.has_discrepancy).label('accurate_counts')
            ).filter(
                and_(
                    CycleCountResult.counted_at >= start_dt,
                    CycleCountResult.counted_at <= end_dt
                )
            ).group_by(
                CycleCountResult.counted_by
            ).all()

            # Convert to list for easier handling
            user_performance = []
            for perf in user_performance_raw:
                user_performance.append({
                    'counted_by': perf.counted_by,
                    'total_counts': perf.total_counts,
                    'accurate_counts': perf.accurate_counts or 0,
                    'avg_time_to_count': 0  # Simplified for now
                })

            # Coverage analysis
            total_tools = Tool.query.count()
            total_chemicals = Chemical.query.count()

            counted_tools = db.session.query(
                func.count(func.distinct(CycleCountItem.item_id))
            ).filter(
                and_(
                    CycleCountItem.item_type == 'tool',
                    CycleCountItem.status == 'counted',
                    CycleCountItem.created_at >= start_dt
                )
            ).scalar()

            counted_chemicals = db.session.query(
                func.count(func.distinct(CycleCountItem.item_id))
            ).filter(
                and_(
                    CycleCountItem.item_type == 'chemical',
                    CycleCountItem.status == 'counted',
                    CycleCountItem.created_at >= start_dt
                )
            ).scalar()

            # Batch completion trends - using simpler approach for SQLite compatibility
            batch_trends_raw = db.session.query(
                func.date(CycleCountBatch.created_at).label('date'),
                func.count(CycleCountBatch.id).label('batches_created'),
                func.count(CycleCountBatch.id).filter(CycleCountBatch.status == 'completed').label('batches_completed')
            ).filter(
                and_(
                    CycleCountBatch.created_at >= start_dt,
                    CycleCountBatch.created_at <= end_dt
                )
            ).group_by(
                func.date(CycleCountBatch.created_at)
            ).order_by(
                func.date(CycleCountBatch.created_at)
            ).all()

            # Convert to list for easier handling
            batch_trends = []
            for trend in batch_trends_raw:
                batch_trends.append({
                    'date': trend.date,
                    'batches_created': trend.batches_created,
                    'batches_completed': trend.batches_completed or 0
                })

            # Format results
            analytics = {
                'accuracy_trends': [
                    {
                        'date': trend['date'] if isinstance(trend['date'], str) else trend['date'].isoformat(),
                        'total_counts': trend['total_counts'],
                        'accurate_counts': trend['accurate_counts'],
                        'accuracy_rate': round((trend['accurate_counts'] / trend['total_counts']) * 100, 2) if trend['total_counts'] > 0 else 0
                    }
                    for trend in accuracy_trends
                ],
                'discrepancy_types': [
                    {
                        'type': disc.discrepancy_type or 'unknown',
                        'count': disc.count
                    }
                    for disc in discrepancy_types
                ],
                'user_performance': [
                    {
                        'user_id': perf['counted_by'],
                        'total_counts': perf['total_counts'],
                        'accurate_counts': perf['accurate_counts'],
                        'accuracy_rate': round((perf['accurate_counts'] / perf['total_counts']) * 100, 2) if perf['total_counts'] > 0 else 0,
                        'avg_time_to_count': perf['avg_time_to_count']
                    }
                    for perf in user_performance
                ],
                'coverage': {
                    'tools': {
                        'total': total_tools,
                        'counted': counted_tools or 0,
                        'coverage_rate': round((counted_tools / total_tools) * 100, 2) if total_tools > 0 else 0
                    },
                    'chemicals': {
                        'total': total_chemicals,
                        'counted': counted_chemicals or 0,
                        'coverage_rate': round((counted_chemicals / total_chemicals) * 100, 2) if total_chemicals > 0 else 0
                    }
                },
                'batch_trends': [
                    {
                        'date': trend['date'] if isinstance(trend['date'], str) else trend['date'].isoformat(),
                        'batches_created': trend['batches_created'],
                        'batches_completed': trend['batches_completed'],
                        'completion_rate': round((trend['batches_completed'] / trend['batches_created']) * 100, 2) if trend['batches_created'] > 0 else 0
                    }
                    for trend in batch_trends
                ]
            }

            return jsonify(analytics), 200

        except Exception as e:
            print(f"Error getting cycle count analytics: {str(e)}")
            return jsonify({'error': 'An error occurred while fetching cycle count analytics'}), 500

    # Export cycle count schedule data
    @app.route('/api/cycle-counts/schedules/<int:schedule_id>/export', methods=['GET'])
    @tool_manager_required
    def export_cycle_count_schedule(schedule_id):
        try:
            from io import StringIO
            import csv
            from flask import make_response

            # Get format parameter
            format_type = request.args.get('format', 'csv').lower()

            # Get schedule
            schedule = CycleCountSchedule.query.get_or_404(schedule_id)

            # Get all batches for the schedule
            batches = CycleCountBatch.query.filter_by(schedule_id=schedule_id).all()

            if format_type == 'excel':
                # Excel export using openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                import io

                wb = Workbook()
                ws = wb.active
                ws.title = "Schedule Details"

                # Add schedule information
                ws['A1'] = "Cycle Count Schedule Export"
                ws['A1'].font = Font(size=16, bold=True)

                ws['A3'] = "Schedule Name:"
                ws['B3'] = schedule.name
                ws['A4'] = "Description:"
                ws['B4'] = schedule.description or ''
                ws['A5'] = "Frequency:"
                ws['B5'] = schedule.frequency
                ws['A6'] = "Method:"
                ws['B6'] = schedule.method
                ws['A7'] = "Created:"
                ws['B7'] = schedule.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ws['A8'] = "Active:"
                ws['B8'] = 'Yes' if schedule.is_active else 'No'

                # Add batch information
                ws['A10'] = "Associated Batches"
                ws['A10'].font = Font(bold=True)

                # Headers for batch data
                headers = ['Batch ID', 'Batch Name', 'Status', 'Created Date', 'Start Date', 'End Date', 'Progress']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=11, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                # Add batch data
                for row, batch in enumerate(batches, 12):
                    ws.cell(row=row, column=1, value=batch.id)
                    ws.cell(row=row, column=2, value=batch.name)
                    ws.cell(row=row, column=3, value=batch.status)
                    ws.cell(row=row, column=4, value=batch.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=row, column=5, value=batch.start_date.strftime('%Y-%m-%d %H:%M:%S') if batch.start_date else '')
                    ws.cell(row=row, column=6, value=batch.end_date.strftime('%Y-%m-%d %H:%M:%S') if batch.end_date else '')

                    # Calculate progress
                    total_items = CycleCountItem.query.filter_by(batch_id=batch.id).count()
                    counted_items = CycleCountItem.query.filter_by(batch_id=batch.id, status='counted').count()
                    progress = f"{counted_items}/{total_items} ({round((counted_items/total_items)*100, 1) if total_items > 0 else 0}%)"
                    ws.cell(row=row, column=7, value=progress)

                # Save to buffer
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                # Create response
                response = make_response(buffer.getvalue())
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response.headers['Content-Disposition'] = f'attachment; filename=cycle_count_schedule_{schedule_id}.xlsx'

                return response

            else:
                # CSV export
                output = StringIO()
                writer = csv.writer(output)

                # Write schedule information
                writer.writerow(['Cycle Count Schedule Export'])
                writer.writerow([])
                writer.writerow(['Schedule Name', schedule.name])
                writer.writerow(['Description', schedule.description or ''])
                writer.writerow(['Frequency', schedule.frequency])
                writer.writerow(['Method', schedule.method])
                writer.writerow(['Created', schedule.created_at.strftime('%Y-%m-%d %H:%M:%S')])
                writer.writerow(['Active', 'Yes' if schedule.is_active else 'No'])
                writer.writerow([])

                # Write batch information
                writer.writerow(['Associated Batches'])
                writer.writerow(['Batch ID', 'Batch Name', 'Status', 'Created Date', 'Start Date', 'End Date', 'Progress'])

                for batch in batches:
                    # Calculate progress
                    total_items = CycleCountItem.query.filter_by(batch_id=batch.id).count()
                    counted_items = CycleCountItem.query.filter_by(batch_id=batch.id, status='counted').count()
                    progress = f"{counted_items}/{total_items} ({round((counted_items/total_items)*100, 1) if total_items > 0 else 0}%)"

                    writer.writerow([
                        batch.id,
                        batch.name,
                        batch.status,
                        batch.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                        batch.start_date.strftime('%Y-%m-%d %H:%M:%S') if batch.start_date else '',
                        batch.end_date.strftime('%Y-%m-%d %H:%M:%S') if batch.end_date else '',
                        progress
                    ])

                # Create response
                response = make_response(output.getvalue())
                response.headers['Content-Type'] = 'text/csv'
                response.headers['Content-Disposition'] = f'attachment; filename=cycle_count_schedule_{schedule_id}.csv'

                return response

        except Exception as e:
            print(f"Error exporting cycle count schedule {schedule_id}: {str(e)}")
            return jsonify({'error': f'An error occurred while exporting cycle count schedule {schedule_id}'}), 500

    # Export cycle count results with discrepancy information
    @app.route('/api/cycle-counts/results/export', methods=['GET'])
    @tool_manager_required
    def export_cycle_count_results():
        try:
            from io import StringIO
            import csv
            from flask import make_response
            from sqlalchemy import and_

            # Get query parameters
            format_type = request.args.get('format', 'csv').lower()
            batch_id = request.args.get('batch_id')
            schedule_id = request.args.get('schedule_id')
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            discrepancies_only = request.args.get('discrepancies_only', 'false').lower() == 'true'

            # Build query for results
            query = db.session.query(
                CycleCountResult,
                CycleCountItem,
                CycleCountBatch
            ).join(
                CycleCountItem, CycleCountResult.item_id == CycleCountItem.id
            ).join(
                CycleCountBatch, CycleCountItem.batch_id == CycleCountBatch.id
            )

            # Apply filters
            if batch_id:
                query = query.filter(CycleCountBatch.id == batch_id)
            if schedule_id:
                query = query.filter(CycleCountBatch.schedule_id == schedule_id)
            if start_date:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(CycleCountResult.counted_at >= start_dt)
            if end_date:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                query = query.filter(CycleCountResult.counted_at <= end_dt)
            if discrepancies_only:
                query = query.filter(CycleCountResult.has_discrepancy)

            # Execute query
            results = query.order_by(CycleCountResult.counted_at.desc()).all()

            if format_type == 'excel':
                # Excel export using openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                import io

                wb = Workbook()
                ws = wb.active
                ws.title = "Count Results"

                # Add title
                ws['A1'] = "Cycle Count Results Export"
                ws['A1'].font = Font(size=16, bold=True)

                ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws['A4'] = f"Total Results: {len(results)}"
                if discrepancies_only:
                    ws['A5'] = "Filter: Discrepancies Only"

                # Headers
                headers = [
                    'Result ID', 'Batch Name', 'Item Type', 'Item ID', 'Expected Qty',
                    'Actual Qty', 'Expected Location', 'Actual Location', 'Has Discrepancy',
                    'Discrepancy Type', 'Condition', 'Counted By', 'Counted At', 'Notes'
                ]

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=7, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                # Add data
                for row, (result, item, batch) in enumerate(results, 8):
                    ws.cell(row=row, column=1, value=result.id)
                    ws.cell(row=row, column=2, value=batch.name)
                    ws.cell(row=row, column=3, value=item.item_type)
                    ws.cell(row=row, column=4, value=item.item_id)
                    ws.cell(row=row, column=5, value=item.expected_quantity)
                    ws.cell(row=row, column=6, value=result.actual_quantity)
                    ws.cell(row=row, column=7, value=item.expected_location)
                    ws.cell(row=row, column=8, value=result.actual_location)
                    ws.cell(row=row, column=9, value='Yes' if result.has_discrepancy else 'No')
                    ws.cell(row=row, column=10, value=result.discrepancy_type or '')
                    ws.cell(row=row, column=11, value=result.condition or '')
                    ws.cell(row=row, column=12, value=result.counted_by)
                    ws.cell(row=row, column=13, value=result.counted_at.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=row, column=14, value=result.notes or '')

                # Save to buffer
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                # Create response
                response = make_response(buffer.getvalue())
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response.headers['Content-Disposition'] = f'attachment; filename=cycle_count_results_{datetime.now().strftime("%Y%m%d")}.xlsx'

                return response

            else:
                # CSV export
                output = StringIO()
                writer = csv.writer(output)

                # Write header information
                writer.writerow(['Cycle Count Results Export'])
                writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
                writer.writerow([f'Total Results: {len(results)}'])
                if discrepancies_only:
                    writer.writerow(['Filter: Discrepancies Only'])
                writer.writerow([])

                # Write column headers
                writer.writerow([
                    'Result ID', 'Batch Name', 'Item Type', 'Item ID', 'Expected Qty',
                    'Actual Qty', 'Expected Location', 'Actual Location', 'Has Discrepancy',
                    'Discrepancy Type', 'Condition', 'Counted By', 'Counted At', 'Notes'
                ])

                # Write data
                for result, item, batch in results:
                    writer.writerow([
                        result.id,
                        batch.name,
                        item.item_type,
                        item.item_id,
                        item.expected_quantity,
                        result.actual_quantity,
                        item.expected_location,
                        result.actual_location,
                        'Yes' if result.has_discrepancy else 'No',
                        result.discrepancy_type or '',
                        result.condition or '',
                        result.counted_by,
                        result.counted_at.strftime('%Y-%m-%d %H:%M:%S'),
                        result.notes or ''
                    ])

                # Create response
                response = make_response(output.getvalue())
                response.headers['Content-Type'] = 'text/csv'
                response.headers['Content-Disposition'] = f'attachment; filename=cycle_count_results_{datetime.now().strftime("%Y%m%d")}.csv'

                return response

        except Exception as e:
            print(f"Error exporting cycle count results: {str(e)}")
            return jsonify({'error': 'An error occurred while exporting cycle count results'}), 500

    # Import cycle count schedules
    @app.route('/api/cycle-counts/schedules/import', methods=['POST'])
    @tool_manager_required
    def import_cycle_count_schedules():
        try:
            import csv
            from io import StringIO

            # Get uploaded file
            if 'file' not in request.files:
                return jsonify({'error': 'No file uploaded'}), 400

            file = request.files['file']
            if file.filename == '':
                return jsonify({'error': 'No file selected'}), 400

            # Read CSV data
            csv_data = file.read().decode('utf-8')
            csv_reader = csv.DictReader(StringIO(csv_data))

            imported_count = 0
            errors = []

            for row_num, row in enumerate(csv_reader, start=2):
                try:
                    # Validate required fields
                    if not row.get('Schedule Name'):
                        errors.append(f"Row {row_num}: Missing Schedule Name")
                        continue

                    if not row.get('Frequency'):
                        errors.append(f"Row {row_num}: Missing Frequency")
                        continue

                    if not row.get('Method'):
                        errors.append(f"Row {row_num}: Missing Method")
                        continue

                    # Validate frequency
                    valid_frequencies = ['daily', 'weekly', 'monthly', 'quarterly', 'yearly']
                    frequency = row['Frequency'].lower()
                    if frequency not in valid_frequencies:
                        errors.append(f"Row {row_num}: Invalid frequency '{row['Frequency']}'. Must be one of: {', '.join(valid_frequencies)}")
                        continue

                    # Validate method
                    valid_methods = ['ABC', 'random', 'location', 'category']
                    method = row['Method'].lower()
                    if method not in valid_methods:
                        errors.append(f"Row {row_num}: Invalid method '{row['Method']}'. Must be one of: {', '.join(valid_methods)}")
                        continue

                    # Check if schedule already exists
                    existing_schedule = CycleCountSchedule.query.filter_by(name=row['Schedule Name']).first()
                    if existing_schedule:
                        errors.append(f"Row {row_num}: Schedule '{row['Schedule Name']}' already exists")
                        continue

                    # Parse active status
                    is_active = True
                    if row.get('Active'):
                        is_active = row['Active'].lower() in ['yes', 'true', '1', 'active']

                    # Create new schedule
                    schedule = CycleCountSchedule(
                        name=row['Schedule Name'],
                        description=row.get('Description', ''),
                        frequency=frequency,
                        method=method,
                        created_by=session['user_id'],
                        is_active=is_active
                    )

                    # Save to database
                    db.session.add(schedule)
                    imported_count += 1

                except ValueError as ve:
                    errors.append(f"Row {row_num}: Invalid data - {str(ve)}")
                except Exception as e:
                    errors.append(f"Row {row_num}: Error - {str(e)}")

            # Commit changes
            db.session.commit()

            # Log the action
            log = AuditLog(
                action_type='cycle_count_schedules_imported',
                action_details=f"Imported {imported_count} cycle count schedules"
            )
            db.session.add(log)
            db.session.commit()

            return jsonify({
                'message': f'Successfully imported {imported_count} schedules',
                'imported_count': imported_count,
                'errors': errors
            }), 200

        except Exception as e:
            db.session.rollback()
            print(f"Error importing cycle count schedules: {str(e)}")
            return jsonify({'error': 'An error occurred while importing cycle count schedules'}), 500

    # Import cycle count batches
    @app.route('/api/cycle-counts/batches/import', methods=['POST'])
    @tool_manager_required
    def import_cycle_count_batches():
        try:
            import csv
            from io import StringIO

            # Get uploaded file
            if 'file' not in request.files:
                return jsonify({'error': 'No file uploaded'}), 400

            file = request.files['file']
            if file.filename == '':
                return jsonify({'error': 'No file selected'}), 400

            # Read CSV data
            csv_data = file.read().decode('utf-8')
            csv_reader = csv.DictReader(StringIO(csv_data))

            imported_count = 0
            errors = []

            for row_num, row in enumerate(csv_reader, start=2):
                try:
                    # Validate required fields
                    if not row.get('Batch Name'):
                        errors.append(f"Row {row_num}: Missing Batch Name")
                        continue

                    # Check if batch already exists
                    existing_batch = CycleCountBatch.query.filter_by(name=row['Batch Name']).first()
                    if existing_batch:
                        errors.append(f"Row {row_num}: Batch '{row['Batch Name']}' already exists")
                        continue

                    # Get schedule if specified
                    schedule_id = None
                    if row.get('Schedule Name'):
                        schedule = CycleCountSchedule.query.filter_by(name=row['Schedule Name']).first()
                        if schedule:
                            schedule_id = schedule.id
                        else:
                            errors.append(f"Row {row_num}: Schedule '{row['Schedule Name']}' not found")
                            continue

                    # Parse dates
                    start_date = None
                    end_date = None
                    if row.get('Start Date'):
                        try:
                            start_date = datetime.strptime(row['Start Date'], '%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            try:
                                start_date = datetime.strptime(row['Start Date'], '%Y-%m-%d')
                            except ValueError:
                                errors.append(f"Row {row_num}: Invalid start date format. Use YYYY-MM-DD or YYYY-MM-DD HH:MM:SS")
                                continue

                    if row.get('End Date'):
                        try:
                            end_date = datetime.strptime(row['End Date'], '%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            try:
                                end_date = datetime.strptime(row['End Date'], '%Y-%m-%d')
                            except ValueError:
                                errors.append(f"Row {row_num}: Invalid end date format. Use YYYY-MM-DD or YYYY-MM-DD HH:MM:SS")
                                continue

                    # Validate status
                    status = 'pending'
                    if row.get('Status'):
                        valid_statuses = ['pending', 'in_progress', 'completed', 'cancelled']
                        if row['Status'].lower() in valid_statuses:
                            status = row['Status'].lower()
                        else:
                            errors.append(f"Row {row_num}: Invalid status '{row['Status']}'. Must be one of: {', '.join(valid_statuses)}")
                            continue

                    # Create new batch
                    batch = CycleCountBatch(
                        name=row['Batch Name'],
                        description=row.get('Description', ''),
                        schedule_id=schedule_id,
                        status=status,
                        start_date=start_date,
                        end_date=end_date,
                        created_by=session['user_id']
                    )

                    # Save to database
                    db.session.add(batch)
                    imported_count += 1

                except ValueError as ve:
                    errors.append(f"Row {row_num}: Invalid data - {str(ve)}")
                except Exception as e:
                    errors.append(f"Row {row_num}: Error - {str(e)}")

            # Commit changes
            db.session.commit()

            # Log the action
            log = AuditLog(
                action_type='cycle_count_batches_imported',
                action_details=f"Imported {imported_count} cycle count batches"
            )
            db.session.add(log)
            db.session.commit()

            return jsonify({
                'message': f'Successfully imported {imported_count} batches',
                'imported_count': imported_count,
                'errors': errors
            }), 200

        except Exception as e:
            db.session.rollback()
            print(f"Error importing cycle count batches: {str(e)}")
            return jsonify({'error': 'An error occurred while importing cycle count batches'}), 500
