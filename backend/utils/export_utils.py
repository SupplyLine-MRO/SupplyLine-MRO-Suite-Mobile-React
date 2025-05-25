"""
Export utilities for generating PDF and Excel reports
"""
import io
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill


def generate_pdf_report(report_data, report_type, timeframe):
    """Generate PDF report from report data"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []

    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )

    # Add title
    title = get_report_title(report_type, timeframe)
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 12))

    # Add generation date
    date_str = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    story.append(Paragraph(date_str, styles['Normal']))
    story.append(Spacer(1, 20))

    # Add report content based on type
    if report_type == 'tool-inventory':
        add_tool_inventory_pdf_content(story, report_data, styles)
    elif report_type == 'checkout-history':
        add_checkout_history_pdf_content(story, report_data, styles)
    elif report_type == 'department-usage':
        add_department_usage_pdf_content(story, report_data, styles)
    elif report_type.startswith('cycle-count'):
        add_cycle_count_pdf_content(story, report_data, report_type, styles)

    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_excel_report(report_data, report_type, timeframe):
    """Generate Excel report from report data"""
    buffer = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Set title
    title = get_report_title(report_type, timeframe)
    worksheet.title = title[:31]  # Excel sheet names limited to 31 chars

    # Add header
    worksheet['A1'] = title
    worksheet['A1'].font = Font(size=16, bold=True)
    worksheet['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    # Add report content based on type
    if report_type == 'tool-inventory':
        add_tool_inventory_excel_content(worksheet, report_data)
    elif report_type == 'checkout-history':
        add_checkout_history_excel_content(worksheet, report_data)
    elif report_type == 'department-usage':
        add_department_usage_excel_content(worksheet, report_data)
    elif report_type.startswith('cycle-count'):
        add_cycle_count_excel_content(worksheet, report_data, report_type)

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                continue
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def get_report_title(report_type, timeframe):
    """Get formatted report title"""
    titles = {
        'tool-inventory': 'Tool Inventory Report',
        'checkout-history': 'Checkout History Report',
        'department-usage': 'Department Usage Report',
        'cycle-count-accuracy': 'Cycle Count Accuracy Report',
        'cycle-count-discrepancies': 'Cycle Count Discrepancy Report',
        'cycle-count-performance': 'Cycle Count Performance Report',
        'cycle-count-coverage': 'Cycle Count Coverage Report'
    }

    title = titles.get(report_type, 'Report')
    if timeframe and timeframe != 'all':
        title += f" ({timeframe.title()})"

    return title


def add_tool_inventory_pdf_content(story, data, styles):
    """Add tool inventory content to PDF"""
    if not data:
        story.append(Paragraph("No tools found.", styles['Normal']))
        return

    # Create table data
    table_data = [['Tool Number', 'Serial Number', 'Description', 'Category', 'Location', 'Status']]

    for tool in data:
        table_data.append([
            tool.get('tool_number', ''),
            tool.get('serial_number', ''),
            tool.get('description', '')[:30] + '...' if len(tool.get('description', '')) > 30 else tool.get('description', ''),
            tool.get('category', ''),
            tool.get('location', ''),
            tool.get('status', '')
        ])

    # Create table
    table = Table(table_data, colWidths=[1*inch, 1*inch, 2*inch, 1*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    story.append(table)


def add_checkout_history_pdf_content(story, data, styles):
    """Add checkout history content to PDF"""
    checkouts = data.get('checkouts', [])
    stats = data.get('stats', {})

    # Add summary statistics
    story.append(Paragraph("Summary Statistics", styles['Heading2']))
    summary_data = [
        ['Total Checkouts', str(stats.get('totalCheckouts', 0))],
        ['Returned Checkouts', str(stats.get('returnedCheckouts', 0))],
        ['Currently Checked Out', str(stats.get('currentlyCheckedOut', 0))],
        ['Average Duration (days)', str(stats.get('averageDuration', 0))]
    ]

    summary_table = Table(summary_data, colWidths=[2*inch, 1*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold')
    ]))

    story.append(summary_table)
    story.append(Spacer(1, 20))

    # Add checkout details
    if checkouts:
        story.append(Paragraph("Checkout Details", styles['Heading2']))

        table_data = [['Tool Number', 'User', 'Department', 'Checkout Date', 'Return Date', 'Duration (days)']]

        for checkout in checkouts[:50]:  # Limit to first 50 for PDF
            table_data.append([
                checkout.get('tool_number', ''),
                checkout.get('user_name', ''),
                checkout.get('department', ''),
                checkout.get('checkout_date', '')[:10] if checkout.get('checkout_date') else '',
                checkout.get('return_date', '')[:10] if checkout.get('return_date') else 'Active',
                str(checkout.get('duration', ''))
            ])

        checkout_table = Table(table_data, colWidths=[1*inch, 1.2*inch, 1*inch, 1*inch, 1*inch, 0.8*inch])
        checkout_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(checkout_table)


def add_department_usage_pdf_content(story, data, styles):
    """Add department usage content to PDF"""
    departments = data.get('departments', [])

    if not departments:
        story.append(Paragraph("No department usage data found.", styles['Normal']))
        return

    # Create table data
    table_data = [['Department', 'Total Checkouts', 'Currently Checked Out', 'Avg Duration (days)', 'Most Used Category']]

    for dept in departments:
        table_data.append([
            dept.get('name', ''),
            str(dept.get('totalCheckouts', 0)),
            str(dept.get('currentlyCheckedOut', 0)),
            str(dept.get('averageDuration', 0)),
            dept.get('mostUsedCategory', '')
        ])

    # Create table
    table = Table(table_data, colWidths=[1.5*inch, 1*inch, 1.2*inch, 1*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    story.append(table)


def add_cycle_count_pdf_content(story, data, report_type, styles):
    """Add cycle count content to PDF"""
    story.append(Paragraph(f"Cycle Count Report: {report_type}", styles['Heading2']))

    if report_type == 'cycle-count-accuracy':
        summary = data.get('summary', {})
        story.append(Paragraph(f"Total Counts: {summary.get('total_counts', 0)}", styles['Normal']))
        story.append(Paragraph(f"Accurate Counts: {summary.get('accurate_counts', 0)}", styles['Normal']))
        story.append(Paragraph(f"Accuracy Rate: {summary.get('accuracy_rate', 0)}%", styles['Normal']))

    # Add more cycle count specific content as needed
    story.append(Paragraph("Detailed cycle count reporting is available in the web interface.", styles['Normal']))


def add_tool_inventory_excel_content(worksheet, data):
    """Add tool inventory content to Excel"""
    if not data:
        worksheet['A4'] = "No tools found."
        return

    # Headers
    headers = ['Tool Number', 'Serial Number', 'Description', 'Category', 'Location', 'Status', 'Condition']
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Data
    for row, tool in enumerate(data, 5):
        worksheet.cell(row=row, column=1, value=tool.get('tool_number', ''))
        worksheet.cell(row=row, column=2, value=tool.get('serial_number', ''))
        worksheet.cell(row=row, column=3, value=tool.get('description', ''))
        worksheet.cell(row=row, column=4, value=tool.get('category', ''))
        worksheet.cell(row=row, column=5, value=tool.get('location', ''))
        worksheet.cell(row=row, column=6, value=tool.get('status', ''))
        worksheet.cell(row=row, column=7, value=tool.get('condition', ''))


def add_checkout_history_excel_content(worksheet, data):
    """Add checkout history content to Excel"""
    checkouts = data.get('checkouts', [])
    stats = data.get('stats', {})

    # Summary statistics
    worksheet['A4'] = "Summary Statistics"
    worksheet['A4'].font = Font(bold=True)

    worksheet['A5'] = "Total Checkouts"
    worksheet['B5'] = stats.get('totalCheckouts', 0)
    worksheet['A6'] = "Returned Checkouts"
    worksheet['B6'] = stats.get('returnedCheckouts', 0)
    worksheet['A7'] = "Currently Checked Out"
    worksheet['B7'] = stats.get('currentlyCheckedOut', 0)
    worksheet['A8'] = "Average Duration (days)"
    worksheet['B8'] = stats.get('averageDuration', 0)

    # Checkout details
    if checkouts:
        worksheet['A10'] = "Checkout Details"
        worksheet['A10'].font = Font(bold=True)

        headers = ['Tool Number', 'User', 'Department', 'Checkout Date', 'Return Date', 'Duration (days)']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=11, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for row, checkout in enumerate(checkouts, 12):
            worksheet.cell(row=row, column=1, value=checkout.get('tool_number', ''))
            worksheet.cell(row=row, column=2, value=checkout.get('user_name', ''))
            worksheet.cell(row=row, column=3, value=checkout.get('department', ''))
            worksheet.cell(row=row, column=4, value=checkout.get('checkout_date', ''))
            worksheet.cell(row=row, column=5, value=checkout.get('return_date', '') or 'Active')
            worksheet.cell(row=row, column=6, value=checkout.get('duration', ''))


def add_department_usage_excel_content(worksheet, data):
    """Add department usage content to Excel"""
    departments = data.get('departments', [])

    if not departments:
        worksheet['A4'] = "No department usage data found."
        return

    # Headers
    headers = ['Department', 'Total Checkouts', 'Currently Checked Out', 'Avg Duration (days)', 'Most Used Category']
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Data
    for row, dept in enumerate(departments, 5):
        worksheet.cell(row=row, column=1, value=dept.get('name', ''))
        worksheet.cell(row=row, column=2, value=dept.get('totalCheckouts', 0))
        worksheet.cell(row=row, column=3, value=dept.get('currentlyCheckedOut', 0))
        worksheet.cell(row=row, column=4, value=dept.get('averageDuration', 0))
        worksheet.cell(row=row, column=5, value=dept.get('mostUsedCategory', ''))


def add_cycle_count_excel_content(worksheet, data, report_type):
    """Add cycle count content to Excel"""
    worksheet['A4'] = f"Cycle Count Report: {report_type}"
    worksheet['A4'].font = Font(bold=True)

    if report_type == 'cycle-count-accuracy':
        summary = data.get('summary', {})
        worksheet['A6'] = "Total Counts"
        worksheet['B6'] = summary.get('total_counts', 0)
        worksheet['A7'] = "Accurate Counts"
        worksheet['B7'] = summary.get('accurate_counts', 0)
        worksheet['A8'] = "Accuracy Rate (%)"
        worksheet['B8'] = summary.get('accuracy_rate', 0)

    # Add more cycle count specific content as needed
    worksheet['A10'] = "Detailed cycle count reporting is available in the web interface."
